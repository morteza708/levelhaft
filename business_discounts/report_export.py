"""
خروجی Excel از گزارش تخفیف‌های بیزنس (با همان فیلترهای صفحه گزارش).
فقط export — بدون import.
"""
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from config.jalali import format_jalali_datetime
from .report_services import (
    apply_report_filters,
    build_code_summary,
    build_representative_summary,
    build_totals,
    get_valid_report_usages,
)


def _style_header_row(ws, row=1):
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='166534', end_color='166534', fill_type='solid')
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill


def _autosize_columns(ws, sample_limit=80):
    """عرض ستون از هدر + نمونهٔ اول ردیف‌ها (سریع‌تر برای شیت‌های بزرگ)."""
    for col in ws.columns:
        column_letter = col[0].column_letter
        max_length = 0
        for cell in col[:1]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        for cell in col[1:1 + sample_limit]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)


def _write_sheet(ws, headers, rows):
    ws.append(headers)
    _style_header_row(ws)
    for row in rows:
        ws.append(row)
    sample_limit = 50 if len(rows) > 200 else 200
    _autosize_columns(ws, sample_limit=sample_limit)


def _user_type_fa(user):
    return 'بیوتیشن' if getattr(user, 'is_beautician', False) else 'عادی'


def _payment_fields(order):
    payments = list(order.payments.all())
    if not payments:
        return '', '', 0

    methods = []
    tx_ids = []
    paid_total = 0
    for payment in payments:
        status_label = payment.get_status_display()
        methods.append(
            f'{payment.get_payment_type_display()}: {payment.amount:,} ({status_label})'
        )
        if payment.transaction_id:
            tx_ids.append(payment.transaction_id)
        if payment.status == 'completed':
            paid_total += payment.amount

    return ' | '.join(methods), ' | '.join(tx_ids), paid_total


def _build_order_detail_row(usage):
    order = usage.order
    discount = usage.discount
    user = usage.user
    payment_methods, payment_tx_ids, paid_total = _payment_fields(order)

    return [
        format_jalali_datetime(order.created_at, fmt='%Y/%m/%d %H:%M'),
        format_jalali_datetime(usage.created_at, fmt='%Y/%m/%d %H:%M'),
        order.order_number,
        user.phone_number,
        user.get_full_name() or '',
        _user_type_fa(user),
        discount.code,
        discount.title,
        discount.discount_type_display,
        discount.sales_representative,
        order.total_amount,
        usage.amount,
        order.discount_amount,
        order.final_amount,
        order.unpaid_amount,
        order.get_status_display(),
        order.get_payment_status_display(),
        order.pasargad_url_id or '',
        payment_methods,
        payment_tx_ids,
        paid_total,
        order.receiver_name,
        order.receiver_phone,
        order.receiver_city,
        order.receiver_postal_code,
    ]


DETAIL_HEADERS = [
    'تاریخ سفارش',
    'تاریخ مصرف کد',
    'شماره سفارش',
    'موبایل مشتری',
    'نام مشتری',
    'نوع کاربر',
    'کد تخفیف',
    'عنوان تخفیف',
    'نوع تخفیف',
    'نماینده',
    'مبلغ قبل تخفیف',
    'مبلغ تخفیف (کد)',
    'مبلغ تخفیف (سفارش)',
    'مبلغ نهایی',
    'مبلغ پرداخت‌نشده',
    'وضعیت سفارش',
    'وضعیت پرداخت',
    'شناسه پاسارگاد',
    'روش‌های پرداخت',
    'شناسه تراکنش‌ها',
    'جمع پرداخت‌شده',
    'نام گیرنده',
    'تلفن گیرنده',
    'شهر',
    'کد پستی',
]


def build_discount_report_excel(request):
    base_qs = get_valid_report_usages()
    filtered_qs = apply_report_filters(base_qs, request.GET)

    totals = build_totals(filtered_qs)
    rep_summary = build_representative_summary(filtered_qs)
    code_summary = build_code_summary(filtered_qs)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'خلاصه'
    ws_summary.append(['شاخص', 'مقدار'])
    _style_header_row(ws_summary)
    ws_summary.append(['تعداد سفارش', totals['order_count']])
    ws_summary.append(['مشتریان یکتا', totals['customer_count']])
    ws_summary.append(['جمع قبل تخفیف', totals['total_before_discount']])
    ws_summary.append(['جمع تخفیف', totals['total_discount']])
    ws_summary.append(['جمع نهایی', totals['total_final']])
    _autosize_columns(ws_summary, sample_limit=10)

    ws_rep = wb.create_sheet('نماینده‌ها')
    _write_sheet(
        ws_rep,
        [
            'نماینده', 'کدهای استفاده‌شده', 'سفارش', 'مشتری',
            'جمع قبل تخفیف', 'جمع تخفیف', 'جمع نهایی',
        ],
        [
            [
                r['sales_representative'],
                r['codes_used'],
                r['order_count'],
                r['customer_count'],
                r['total_before_discount'],
                r['total_discount'],
                r['total_final'],
            ]
            for r in rep_summary
        ],
    )

    ws_codes = wb.create_sheet('کدهای تخفیف')
    _write_sheet(
        ws_codes,
        [
            'کد', 'عنوان', 'نماینده', 'سفارش', 'مشتری',
            'قبل تخفیف', 'تخفیف', 'نهایی', 'میانگین سفارش',
        ],
        [
            [
                r['code'],
                r['title'],
                r['sales_representative'],
                r['order_count'],
                r['customer_count'],
                r['total_before_discount'],
                r['total_discount'],
                r['total_final'],
                r['avg_final'],
            ]
            for r in code_summary
        ],
    )

    ws_details = wb.create_sheet('جزئیات تراکنش')
    details_qs = filtered_qs.prefetch_related('order__payments').order_by('-order__created_at')
    detail_rows = [_build_order_detail_row(usage) for usage in details_qs]
    _write_sheet(ws_details, DETAIL_HEADERS, detail_rows)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def discount_report_excel_response(request):
    content = build_discount_report_excel(request)
    response = HttpResponse(
        content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="discount-report.xlsx"'
    return response
